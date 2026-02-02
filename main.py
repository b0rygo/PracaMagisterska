from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import time
import os
import io

from models.tesseract import extract_classic
from models.donut import extract_donut
from models.llama import extract_llama

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.post("/upload/{method}")
async def process_invoice(method: str, file: UploadFile = File(...)):
    start_time = time.time()

    # Zapis pliku lokalnie
    upload_path = f"uploads/{file.filename}"
    with open(upload_path, "wb+") as f:
        f.write(file.file.read())

    # Logika wyboru ścieżki badawczej
    try:
        if method == "tesseract":
            result = extract_classic(upload_path)
            result["method_info"] = "Tesseract OCR + Regex"

        elif method == "donut":
            raw_result = extract_donut(upload_path)
            result = {
                "method_info": "DONUT Transformer",
                "extracted_data": raw_result
            }

        elif method == "llama":
            raw_result = extract_llama(upload_path)
            result = {
                "method_info": "OCR + LLaMA3 LLM",
                "extracted_data": raw_result
            }

        else:
            raise HTTPException(status_code=400, detail="Nieznana metoda. Dostępne: tesseract, donut, llama")

    except Exception as e:
        return {"status": "error", "message": str(e)}

    execution_time = time.time() - start_time

    return {
        "method": method,
        "results": result,
        "execution_time": f"{execution_time:.2f}s"
    }


@app.post("/export/xlsx")
async def export_xlsx(data: dict):
    """Eksport danych faktury do formatu XLSX"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Faktura"

        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a2a6c", end_color="1a2a6c", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Nagłówki
        headers = [
            "Sprzedawca - Nazwa", "Sprzedawca - NIP",
            "Nabywca - Nazwa", "Nabywca - NIP",
            "Numer Faktury", "Data Wystawienia", "Data Wykonania Usługi",
            "Kwota Brutto", "Termin Płatności", "Sposób Płatności",
            "Bank", "Nr Konta"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        # Dane
        row_data = [
            data.get('sprzedawca', {}).get('nazwa', ''),
            data.get('sprzedawca', {}).get('nip', ''),
            data.get('nabywca', {}).get('nazwa', ''),
            data.get('nabywca', {}).get('nip', ''),
            data.get('faktura', {}).get('numer', ''),
            data.get('faktura', {}).get('data_wystawienia', ''),
            data.get('faktura', {}).get('data_wykonania_uslugi', ''),
            data.get('platnosc', {}).get('kwota_brutto', ''),
            data.get('platnosc', {}).get('termin_platnosci', ''),
            data.get('platnosc', {}).get('sposob_platnosci', ''),
            data.get('platnosc', {}).get('bank', ''),
            data.get('platnosc', {}).get('nr_konta', '')
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left')

        # Zapisz do bufora
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=faktura_export.xlsx"}
        )

    except ImportError:
        raise HTTPException(status_code=500, detail="Biblioteka openpyxl nie jest zainstalowana. Uruchom: pip install openpyxl")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Błąd eksportu: {str(e)}")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="127.0.0.1", port=8000)